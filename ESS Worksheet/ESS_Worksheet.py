import csv
import openpyxl
import os.path
from os import path

wb = openpyxl.load_workbook('Worksheet.xlsx')
ws = wb.active

with open('Jobs by Estimator.csv') as csvFile:
    reader = csv.reader(csvFile, delimiter='\t')
    lineCount = 0
    for row in reader:
        if lineCount == 0:
            next(reader)
            lineCount += 1
        else:
            saveName = row[0] + "_" + row[1] + ".xlsx"
            dirName = row[0]
            ws['B1'] = row[1]
            ws['B2'] = row[2]
            ws['B3'] = row[0]
            if path.exists(dirName):
                os.chdir(dirName)
                wb.save(filename = saveName)
            else:
                os.mkdir(dirName)
                os.chdir(dirName)
                wb.save(filename = saveName)
